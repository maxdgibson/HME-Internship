import tkinter as tk
from tkinter import messagebox
import openpyxl
from datetime import datetime

# Function to increase font size for all widgets
def increase_font_size(widget, size):
    font = widget['font']
    font_family = font.split(' ')[0]
    widget.configure(font=(font_family, size))

def increase_dropdown_font_size(dropdown, size):
    menu = dropdown['menu']
    for index in range(menu.index('end') + 1):
        menu.entryconfig(index, font=('TkDefaultFont', size))

def insert_data():
    # Get the values from the input fields
    id_number = id_entry.get()
    category = category_var.get()
    product = product_var.get()
    quantity = quantity_entry.get()
    shift = shift_var.get()
    quarter = quarter_var.get()
    comments = comments_entry.get()
    date_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # Confirmation pop-up
    result = messagebox.askquestion('Confirmation', 'Are you sure you want to submit?', icon='warning')
    if result == 'no':
        return

    # Open the Excel file
    wb = openpyxl.load_workbook('C:/Users/mgibson/OneDrive - HM Electronics, Inc/Desktop/PathToExcel/QuantEntries.xlsx')

    # Select the desired sheet
    sheet = wb['Sheet1']  # Replace 'Sheet1' with the name of your sheet

    # Find the next available row in the sheet
    next_row = sheet.max_row + 1

    # Insert the data into the Excel sheet
    sheet.cell(row=next_row, column=1).value = date_time
    sheet.cell(row=next_row, column=3).value = id_number
    sheet.cell(row=next_row, column=5).value = shift
    sheet.cell(row=next_row, column=7).value = quarter
    sheet.cell(row=next_row, column=9).value = category
    sheet.cell(row=next_row, column=11).value = product
    sheet.cell(row=next_row, column=13).value = quantity
    sheet.cell(row=next_row, column=15).value = comments

    # Save the changes to the Excel file
    wb.save('C:/Users/mgibson/OneDrive - HM Electronics, Inc/Desktop/PathToExcel/QuantEntries.xlsx')

    # Show a success message
    messagebox.showinfo('Success', 'Data inserted successfully!')

    # Reset the input fields
    id_entry.delete(0, 'end')
    shift_var.set(shift_options[0])
    quarter_var.set(quarter_options[0])
    category_var.set(category_options[0])
    product_var.set('Select Product')
    quantity_entry.delete(0, 'end')
    comments_entry.delete(0, 'end')

def update_product_options(*args):
    selected_category = category_var.get()

    # Clear the current product options
    product_dropdown['menu'].delete(0, 'end')

    # Determine the product options based on the selected category
    if selected_category == 'Bases':
        product_options = ['BS6X00', 'BS6000', 'BS7X00', 'Console', 'Router', 'Pro Aud']
    elif selected_category == 'PCB':
        product_options = ['PCBHS6', 'PRGPCBAIO', 'PCBHS7', 'PCBCHG', 'PCBBSX', 'PCBIB7']
    elif selected_category == 'Nexeo':
        product_options = ['RT7000', 'IB7000', 'SM7000']
    elif selected_category == 'CU/TSP':
        product_options = ['CU TEST', 'CU50', 'TSP TEST', 'PCBTSP']
    else:
        product_options = []

    # Set the default option
    product_var.set('')
    product_dropdown['menu'].add_command(label='Select Product', command=tk._setit(product_var, ''))

    # Add the product options to the dropdown menu
    for option in product_options:
        product_dropdown['menu'].add_command(label=option, command=tk._setit(product_var, option))

# Create the Tkinter window
window = tk.Tk()
window.title('Data Entry')
window.geometry('490x330')

# Increase the font size for all labels and entry fields
font_size = 16

id_label = tk.Label(window, text='ID Number:')
id_label.pack(side='top')
increase_font_size(id_label, font_size)

id_entry = tk.Entry(window)
id_entry.pack(side='top')
increase_font_size(id_entry, font_size)

shift_quarter_frame = tk.Frame(window)
shift_quarter_frame.pack(side='top')

shift_label = tk.Label(shift_quarter_frame, text='Shift:')
shift_label.pack(side='left')
increase_font_size(shift_label, font_size)

shift_options = ['1', '2']
shift_var = tk.StringVar(window)
shift_var.set(shift_options[0])  # Set the default option
shift_dropdown = tk.OptionMenu(shift_quarter_frame, shift_var, *shift_options)
shift_dropdown.pack(side='left')
increase_font_size(shift_dropdown, font_size)
increase_dropdown_font_size(shift_dropdown, font_size)

quarter_label = tk.Label(shift_quarter_frame, text='Quarter:')
quarter_label.pack(side='left')
increase_font_size(quarter_label, font_size)

quarter_options = ['1', '2', '3', '4']
quarter_var = tk.StringVar(window)
quarter_var.set(quarter_options[0])  # Set the default option
quarter_dropdown = tk.OptionMenu(shift_quarter_frame, quarter_var, *quarter_options)
quarter_dropdown.pack(side='left')
increase_font_size(quarter_dropdown, font_size)
increase_dropdown_font_size(quarter_dropdown, font_size)

category_product_frame = tk.Frame(window)
category_product_frame.pack(side='top')

category_label = tk.Label(category_product_frame, text='Category:')
category_label.pack(side='left')
increase_font_size(category_label, font_size)

category_options = ['Bases', 'PCB', 'Nexeo', 'CU/TSP']
category_var = tk.StringVar(window)
category_var.set(category_options[0])  # Set the default option
category_var.trace('w', update_product_options)
category_dropdown = tk.OptionMenu(category_product_frame, category_var, *category_options)
category_dropdown.pack(side='left', padx=5)
increase_font_size(category_dropdown, font_size)
increase_dropdown_font_size(category_dropdown, font_size)

product_label = tk.Label(category_product_frame, text='Product:')
product_label.pack(side='left')
increase_font_size(product_label, font_size)

product_var = tk.StringVar(window)
product_var.set('Select Product')  # Set the default option
product_dropdown = tk.OptionMenu(category_product_frame, product_var, 'Select Product')
product_dropdown.pack(side='left', padx=5)
increase_font_size(product_dropdown, font_size)
increase_dropdown_font_size(product_dropdown, font_size)

quantity_label = tk.Label(window, text='Quantity:')
quantity_label.pack(side='top')
increase_font_size(quantity_label, font_size)

quantity_entry = tk.Entry(window)
quantity_entry.pack(side='top')
increase_font_size(quantity_entry, font_size)

comments_label = tk.Label(window, text='Comments:')
comments_label.pack(side='top')
increase_font_size(comments_label, font_size)

comments_entry = tk.Entry(window)
comments_entry.pack(side='top')
increase_font_size(comments_entry, font_size)

submit_button = tk.Button(window, text='Submit', command=insert_data)
submit_button.pack(side='top')
increase_font_size(submit_button, font_size)

# Start the Tkinter event loop
window.mainloop()
